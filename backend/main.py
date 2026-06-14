import os
import json
import asyncio
import tempfile
import logging
from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, Request, UploadFile, File, Form, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, PlainTextResponse, RedirectResponse
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from typing import Optional

import brain
import tasks as tasks_module
import history
import dashboard_db
import graphql_handler
import auth
import telegram_bot
from context_manager import ContextManager

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ── Rate limiter ───────────────────────────────────────────────
limiter = Limiter(key_func=get_remote_address)

app = FastAPI()
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# ── CORS — restrict to the configured frontend origin ──────────
FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:3000")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[FRONTEND_URL, "http://localhost:3000", "http://localhost:5173"],
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "Accept"],
    allow_credentials=False,
)

# Lazy-loaded Whisper model
_whisper_model = None

def _get_whisper():
    global _whisper_model
    if _whisper_model is None:
        from faster_whisper import WhisperModel
        logging.getLogger("faster_whisper").setLevel(logging.WARNING)
        _whisper_model = WhisperModel("tiny", device="cpu", compute_type="int8")
    return _whisper_model


# ── Text extraction helpers ────────────────────────────────────

def _extract_pdf(path: str) -> str:
    import fitz  # pymupdf
    doc = fitz.open(path)
    pages = [page.get_text() for page in doc]
    doc.close()
    return "\n\n".join(p.strip() for p in pages if p.strip())


def _extract_docx(path: str) -> str:
    from docx import Document
    doc = Document(path)
    return "\n\n".join(p.text.strip() for p in doc.paragraphs if p.text.strip())


def _extract_xlsx(path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sections = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append("\t".join(cells))
        if rows:
            sections.append(f"[Sheet: {sheet}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(sections)


def _extract_txt(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


EXTRACTORS = {
    ".pdf":  _extract_pdf,
    ".docx": _extract_docx,
    ".doc":  _extract_docx,
    ".xlsx": _extract_xlsx,
    ".xls":  _extract_xlsx,
    ".txt":  _extract_txt,
    ".md":   _extract_txt,
    ".csv":  _extract_txt,
}


# ── Startup ────────────────────────────────────────────────────

@app.on_event("startup")
async def startup():
    import time
    for attempt in range(10):
        try:
            brain.ensure_collections()
            tasks_module.ensure_collections()
            logger.info("Qdrant collections ready")
            break
        except Exception as e:
            if attempt < 9:
                logger.warning("Qdrant not ready (%d/10): %s — retrying in 3s", attempt + 1, e)
                time.sleep(3)
            else:
                logger.error("Qdrant unreachable after 10 attempts: %s", e)
    history.ensure_db()
    auth.ensure_users_table()
    telegram_bot.ensure_telegram_table()
    dashboard_db.ensure_dashboard_tables()
    try:
        import vault as vault_module
        vault_module.ensure_collection()
        logger.info("Vault collection ready")
    except Exception as e:
        logger.warning("Vault init skipped (VAULT_SECRET not set or Qdrant unavailable): %s", e)
    telegram_bot.start_scheduler()
    _register_dashboard_pipelines()
    if os.getenv("TELEGRAM_BOT_TOKEN") and os.getenv("BACKEND_URL"):
        await telegram_bot.set_webhook(BACKEND_URL)
    # Schedule pipeline run 2 seconds after startup so FastAPI is fully
    # initialised before the pipelines start hitting external APIs.
    loop = asyncio.get_event_loop()
    loop.call_later(2, lambda: asyncio.ensure_future(_run_startup_pipelines()))


# ── Routes ─────────────────────────────────────────────────────

BACKEND_URL = os.getenv("BACKEND_URL", "http://localhost:8000")
FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:3000")


async def _run_startup_pipelines() -> None:
    """Run all dashboard pipelines concurrently (background task).
    Each pipeline is independent and idempotent — safe to parallelise.
    Pipelines use synchronous requests calls internally, so each is run
    in a thread via asyncio.to_thread to avoid blocking the event loop."""
    from pipelines import news, learning, repos, special, transit
    import asyncio as _asyncio

    async def _run(name: str, fn) -> None:
        try:
            logger.info("Pipeline start: %s", name)
            # Run the coroutine — pipelines use requests internally which blocks,
            # but they are async functions so we await them directly. The blocking
            # requests calls are short enough (< 15s timeout each) not to starve
            # other coroutines given they run concurrently via gather.
            await fn()
            logger.info("Pipeline done:  %s", name)
        except Exception as e:
            logger.error("Pipeline error: %s — %s", name, e, exc_info=True)

    logger.info("Starting all dashboard pipelines concurrently")
    await asyncio.gather(
        _run("transit",  lambda: transit.run_pipeline(force=True)),
        _run("special",  special.run_pipeline),
        _run("repos",    repos.run_pipeline),
        _run("news",     news.run_pipeline),
        _run("learning", learning.run_pipeline),
    )
    logger.info("All dashboard pipelines completed")


def _register_dashboard_pipelines() -> None:
    """Register all dashboard background pipelines with the existing APScheduler instance.

    Schedule:
      - news, learning, repos, special — daily at 05:00 local time
        (fresh content ready before the user opens the dashboard in the morning)
      - transit — every 10 minutes between 05:00 and 23:00
    """
    from apscheduler.triggers.cron import CronTrigger
    from apscheduler.triggers.interval import IntervalTrigger
    from telegram_bot import scheduler
    from pipelines import news, learning, repos, special, transit

    def _add(job_fn, trigger, job_id: str) -> None:
        try:
            scheduler.add_job(
                job_fn, trigger, id=job_id,
                replace_existing=True, misfire_grace_time=600,
            )
            logger.info("Dashboard pipeline registered: %s", job_id)
        except Exception as e:
            logger.error("Failed to register pipeline %s: %s", job_id, e)

    # All content pipelines triggered together at 05:00 via a single wrapper job
    # that runs them all concurrently with asyncio.gather
    _add(_run_startup_pipelines, CronTrigger(hour=5, minute=0),  "dashboard_daily_05h")

    # Transit also polls every 10 min during operating hours (05:00–23:00)
    _add(transit.run_pipeline,   IntervalTrigger(minutes=10),    "dashboard_transit_poll")


@app.get("/health")
async def health():
    return {"status": "ok"}


@app.get("/debug/env")
async def debug_env():
    import os
    redirect_uri = f"{os.getenv('BACKEND_URL', '')}/auth/google/callback"
    oauth_url = auth.google_auth_url(redirect_uri)
    return {
        "GOOGLE_CLIENT_ID_set":     bool(os.getenv("GOOGLE_CLIENT_ID")),
        "GOOGLE_CLIENT_SECRET_set": bool(os.getenv("GOOGLE_CLIENT_SECRET")),
        "JWT_SECRET_set":           bool(os.getenv("JWT_SECRET")),
        "BACKEND_URL":              os.getenv("BACKEND_URL", "NOT SET"),
        "FRONTEND_URL":             os.getenv("FRONTEND_URL", "NOT SET"),
        "oauth_url_preview":        oauth_url[:120],
    }


# ── Auth routes ────────────────────────────────────────────────

@app.get("/auth/google")
async def google_login():
    """Redirect browser to Google OAuth consent screen."""
    redirect_uri = f"{BACKEND_URL}/auth/google/callback"
    url = auth.google_auth_url(redirect_uri)
    return RedirectResponse(url)


@app.get("/auth/google/callback")
async def google_callback(request: Request, code: Optional[str] = None, state: Optional[str] = None, error: Optional[str] = None):
    """Handle Google OAuth callback, issue JWT, redirect to frontend."""
    logger.info("OAuth callback — error=%s code_present=%s params=%s", error, bool(code), dict(request.query_params))
    if error:
        logger.error("Google returned error: %s", error)
        return RedirectResponse(f"{FRONTEND_URL}?error={error}")
    if not code:
        logger.error("No code in callback. Params: %s", dict(request.query_params))
        return RedirectResponse(f"{FRONTEND_URL}?error=no_code")
    try:
        redirect_uri = f"{BACKEND_URL}/auth/google/callback"
        logger.info("Exchanging code with redirect_uri=%s", redirect_uri)
        info  = auth.exchange_code(code, redirect_uri)
        logger.info("Got user info: google_id present=%s", bool(info.get("sub")))
        user  = auth.upsert_user(info["google_id"], info["email"], info["name"], info["avatar_url"])
        token = auth.issue_jwt(user)
        logger.info("Issuing JWT for user=%s, redirecting to frontend", user["id"])
        # Trigger dashboard pipelines in the background on every login.
        # Each pipeline is idempotent — if today's cycle already exists it returns in < 1ms.
        asyncio.create_task(_run_startup_pipelines())
        return RedirectResponse(f"{FRONTEND_URL}?token={token}")
    except Exception as e:
        logger.error("OAuth callback error: %s", e, exc_info=True)
        return RedirectResponse(f"{FRONTEND_URL}?error=auth_failed")


@app.get("/auth/me")
async def get_me(authorization: Optional[str] = Header(None)):
    """Return the current user from JWT, including Telegram link status."""
    user_id = auth.extract_user_id(authorization)
    if not user_id:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    user = auth.get_user_by_id(user_id)
    if not user:
        return JSONResponse({"error": "User not found"}, status_code=404)
    telegram_status = telegram_bot.get_telegram_link_status(user_id)
    return JSONResponse({
        "id":             user["id"],
        "email":          user["email"],
        "name":           user["name"],
        "avatar_url":     user["avatar_url"],
        "telegram":       telegram_status,
    })


@app.post("/telegram/link")
async def link_telegram(request: Request, authorization: Optional[str] = Header(None)):
    """
    Link a Telegram account to the currently authenticated Google user.
    The user sends their Telegram numeric ID (from @userinfobot).
    After linking, Telegram messages from that ID are routed to this user's data.
    """
    user_id = auth.extract_user_id(authorization)
    if not user_id:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    body = await request.json()
    telegram_id = str(body.get("telegram_id", "")).strip()
    if not telegram_id:
        return JSONResponse({"error": "telegram_id is required"}, status_code=400)
    ok = telegram_bot.link_telegram_user(telegram_id, user_id)
    if ok:
        return JSONResponse({"linked": True, "telegram_id": telegram_id})
    return JSONResponse({"error": "Failed to link Telegram account"}, status_code=500)


@app.delete("/telegram/link")
async def unlink_telegram(authorization: Optional[str] = Header(None)):
    """Unlink the Telegram account from the current Google user."""
    user_id = auth.extract_user_id(authorization)
    if not user_id:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    status = telegram_bot.get_telegram_link_status(user_id)
    if not status["linked"]:
        return JSONResponse({"error": "No Telegram account linked"}, status_code=404)
    ok = telegram_bot.unlink_telegram_user(user_id)
    if ok:
        return JSONResponse({"unlinked": True})
    return JSONResponse({"error": "Failed to unlink"}, status_code=500)


# ── GraphQL ────────────────────────────────────────────────────

@app.post("/graphql")
@limiter.limit("60/minute")           # 60 requests/min per IP — covers normal use
async def graphql_post(request: Request):
    return await graphql_handler.handle(request)


@app.get("/graphql")
async def graphql_schema():
    return PlainTextResponse(graphql_handler.SCHEMA_SDL, media_type="text/plain")


@app.get("/context-debug")
async def context_debug(message: str = "test", authorization: Optional[str] = Header(None)):
    user_id = auth.extract_user_id(authorization) or "debug"
    ctx = await ContextManager.build(current_message=message, user_id=user_id)
    return PlainTextResponse(ctx.debug(), media_type="text/plain")


# ── Auth logout ────────────────────────────────────────────────

@app.post("/auth/logout")
async def logout(authorization: Optional[str] = Header(None)):
    """Revoke the current JWT immediately. Token is added to revocation list."""
    if not authorization or not authorization.startswith("Bearer "):
        return JSONResponse({"error": "No token provided"}, status_code=400)
    token = authorization[7:].strip()
    ok = auth.revoke_token(token)
    if ok:
        return JSONResponse({"logged_out": True})
    return JSONResponse({"error": "Token invalid or already revoked"}, status_code=400)


@app.post("/transcribe")
@limiter.limit("10/minute")           # Whisper is CPU-heavy — 10/min is generous
async def transcribe(
    request: Request,
    audio: UploadFile = File(...),
    authorization: Optional[str] = Header(None),
):
    """Transcribe audio using faster-whisper. Requires auth."""
    if not auth.extract_user_id(authorization):
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    # Cap upload size at 25MB
    MAX_SIZE = 25 * 1024 * 1024
    content = await audio.read()
    if len(content) > MAX_SIZE:
        return JSONResponse({"error": "Audio file too large (max 25MB)"}, status_code=413)
    try:
        suffix = os.path.splitext(audio.filename or "audio.webm")[1] or ".webm"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)   # use already-read content, not audio.read() again
            tmp_path = tmp.name

        model = _get_whisper()
        segments, _ = model.transcribe(tmp_path, beam_size=5, language="en")
        text = " ".join(seg.text.strip() for seg in segments).strip()
        os.unlink(tmp_path)

        if not text:
            return JSONResponse({"error": "No speech detected"}, status_code=400)
        return JSONResponse({"text": text})

    except Exception as e:
        logger.error("Transcription error: %s", e, exc_info=True)
        return JSONResponse({"error": str(e)}, status_code=500)


# ── Telegram webhook ───────────────────────────────────────────

TELEGRAM_WEBHOOK_SECRET = os.getenv("TELEGRAM_WEBHOOK_SECRET", "")

@app.post("/telegram/webhook")
async def telegram_webhook(request: Request):
    """Receive updates from Telegram and process them.

    Validates X-Telegram-Bot-Api-Secret-Token header when TELEGRAM_WEBHOOK_SECRET is set.
    Never logs message content — only structural metadata.
    """
    # Validate webhook secret token if configured
    if TELEGRAM_WEBHOOK_SECRET:
        incoming_secret = request.headers.get("X-Telegram-Bot-Api-Secret-Token", "")
        if not incoming_secret or incoming_secret != TELEGRAM_WEBHOOK_SECRET:
            logger.warning("Telegram webhook: invalid secret token — rejecting")
            return JSONResponse({"ok": False}, status_code=403)

    body = await request.body()
    # Log only structural info — never message content
    logger.debug("Telegram webhook received: %d bytes", len(body))
    try:
        update = json.loads(body)
        update_id = update.get("update_id", "?")
        update_type = next((k for k in ("message", "edited_message", "callback_query") if k in update), "unknown")
        logger.info("Telegram update id=%s type=%s", update_id, update_type)
        await telegram_bot._handle_update_safe(update)
    except Exception as e:
        logger.error("Telegram webhook error: %s", e, exc_info=True)
    return JSONResponse({"ok": True})


@app.get("/notifications")
async def get_notifications(
    unread_only: bool = True,
    limit: int = 20,
    authorization: Optional[str] = Header(None),
):
    """Return notifications for the authenticated user."""
    user_id = auth.extract_user_id(authorization)
    if not user_id:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    items = history.get_notifications(user_id, unread_only=unread_only, limit=limit)
    return JSONResponse({"notifications": items, "unread_count": len([n for n in items if not n["read"]])})


@app.post("/notifications/read")
async def mark_notifications_read(
    request: Request,
    authorization: Optional[str] = Header(None),
):
    """Mark notifications as read. Pass {ids: [...]} to mark specific ones, or empty body for all."""
    user_id = auth.extract_user_id(authorization)
    if not user_id:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    ids = body.get("ids", None)
    history.mark_notifications_read(user_id, ids)
    return JSONResponse({"ok": True})


@app.get("/telegram/debug")
async def telegram_debug():
    """Debug endpoint — shows token status, allowed IDs, and webhook info."""
    allowed_raw = os.getenv("TELEGRAM_ALLOWED_IDS", "NOT SET")
    token_set   = bool(os.getenv("TELEGRAM_BOT_TOKEN"))
    info        = await telegram_bot._tg_post("getWebhookInfo", {})
    return JSONResponse({
        "token_set":       token_set,
        "allowed_ids_raw": repr(allowed_raw),
        "webhook_info":    info.get("result", {}),
    })


@app.delete("/memory/{point_id}")
async def delete_memory(point_id: str, authorization: Optional[str] = Header(None)):
    """Delete a single memory point scoped to the authenticated user."""
    user_id = auth.extract_user_id(authorization)
    if not user_id:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    ok = brain.delete_memory(point_id, user_id)
    if ok:
        return JSONResponse({"deleted": point_id})
    return JSONResponse({"error": f"Failed to delete {point_id}"}, status_code=500)


@app.delete("/memory/source/{source_id}")
async def delete_memory_source(source_id: str, authorization: Optional[str] = Header(None)):
    user_id = auth.extract_user_id(authorization)
    if not user_id:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    brain.delete_memories_by_source(source_id, user_id)
    return JSONResponse({"deleted_source": source_id})


@app.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    category: str = Form("General"),
    authorization: Optional[str] = Header(None),
):
    """Accept a file, extract text, chunk and save scoped to the authenticated user."""
    user_id = auth.extract_user_id(authorization)
    if not user_id:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    filename = file.filename or "upload"
    ext = os.path.splitext(filename)[1].lower()

    if ext not in EXTRACTORS:
        return JSONResponse(
            {"error": f"Unsupported file type '{ext}'. Supported: {', '.join(EXTRACTORS)}"},
            status_code=400,
        )

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            tmp.write(await file.read())
            tmp_path = tmp.name

        text = EXTRACTORS[ext](tmp_path)
        os.unlink(tmp_path)

        if not text.strip():
            return JSONResponse({"error": "No text could be extracted from the file."}, status_code=400)

        logger.info("Extracted %d chars from %s → category=%s user=%s", len(text), filename, category, user_id[:8])

        result = brain.chunk_and_save(text, category, user_id)

        return JSONResponse({
            "filename":    filename,
            "category":    category,
            "saved":       result["saved_count"],
            "skipped":     result["skipped_count"],
            "total_chars": len(text),
        })

    except Exception as e:
        logger.error("Upload error for %s: %s", filename, e, exc_info=True)
        return JSONResponse({"error": str(e)}, status_code=500)
